#!/usr/bin/env python
"""
Script para importar gastos desde Excel
Compatible con formato: Fecha | Concepto | Categoría | Monto | Tarjeta | Participante(s)

Uso:
  python import_excel.py mi_gastos.xlsx
"""
import sys
import openpyxl
from datetime import datetime
from decimal import Decimal
from sqlalchemy.orm import Session

from app.database import SessionLocal
from app.models.expense import Expense
from app.models.expense_participant import ExpenseParticipant
from app.models.card import Card
from app.models.category import Category
from app.models.user import User

class Color:
    GREEN = '\033[92m'
    YELLOW = '\033[93m'
    RED = '\033[91m'
    ENDC = '\033[0m'

def print_success(msg):
    print(f"{Color.GREEN}✅ {msg}{Color.ENDC}")

def print_warning(msg):
    print(f"{Color.YELLOW}⚠️  {msg}{Color.ENDC}")

def print_error(msg):
    print(f"{Color.RED}❌ {msg}{Color.ENDC}")

def import_from_excel(filepath: str):
    """Importar gastos desde Excel"""
    
    print(f"\n📂 Leyendo archivo: {filepath}")
    
    try:
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active
        print_success(f"Archivo cargado: {sheet.title}")
    except Exception as e:
        print_error(f"No se pudo abrir archivo: {e}")
        return
    
    db: Session = SessionLocal()
    
    try:
        # Obtener el usuario (asumimos que es el primero activo)
        user = db.query(User).filter(User.is_active == True).first()
        if not user:
            print_error("No hay usuarios en la BD")
            return
        
        print_success(f"Usuario: {user.username}")
        
        # Leer filas (saltar header)
        imported = 0
        errors = 0
        
        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Estructura esperada:
                # [Fecha, Concepto, Categoría, Monto, Tarjeta, Participante1, Participante2, ...]
                
                if not any(row):  # Fila vacía
                    continue
                
                fecha = row[0]
                concepto = row[1]
                categoria_name = row[2]
                monto = row[3]
                tarjeta_name = row[4]
                
                # Validar datos obligatorios
                if not all([fecha, concepto, categoria_name, monto, tarjeta_name]):
                    print_warning(f"Fila {idx}: Datos incompletos, saltando")
                    errors += 1
                    continue
                
                # Convertir fecha
                if isinstance(fecha, str):
                    fecha = datetime.strptime(fecha, "%d/%m/%Y")
                elif isinstance(fecha, datetime):
                    pass
                else:
                    print_warning(f"Fila {idx}: Fecha inválida {fecha}")
                    errors += 1
                    continue
                
                # Obtener tarjeta
                card = db.query(Card).filter(
                    Card.user_id == user.id,
                    Card.name == tarjeta_name
                ).first()
                
                if not card:
                    print_warning(f"Fila {idx}: Tarjeta '{tarjeta_name}' no existe")
                    errors += 1
                    continue
                
                # Obtener categoría
                category = db.query(Category).filter(
                    Category.user_id == user.id,
                    Category.name == categoria_name
                ).first()
                
                if not category:
                    print_warning(f"Fila {idx}: Categoría '{categoria_name}' no existe")
                    errors += 1
                    continue
                
                # Convertir monto
                try:
                    monto = Decimal(str(monto))
                except:
                    print_warning(f"Fila {idx}: Monto inválido {monto}")
                    errors += 1
                    continue
                
                # Crear gasto
                expense = Expense(
                    card_id=card.id,
                    category_id=category.id,
                    date=fecha,
                    concept=concepto,
                    total_amount=monto,
                    notes=f"Importado de Excel en {datetime.now().strftime('%Y-%m-%d')}"
                )
                db.add(expense)
                db.flush()
                
                # Participantes (se asume que si solo hay el usuario, es suyo)
                # Si hay más participantes, dividir equitativamente
                participants_data = [p for p in row[5:] if p]
                
                if not participants_data:
                    # Solo el usuario
                    participant = ExpenseParticipant(
                        expense_id=expense.id,
                        user_id=user.id,
                        amount=monto
                    )
                    db.add(participant)
                else:
                    # Dividir entre participantes
                    participant_amount = monto / (len(participants_data) + 1)
                    
                    # El usuario
                    p1 = ExpenseParticipant(
                        expense_id=expense.id,
                        user_id=user.id,
                        amount=participant_amount
                    )
                    db.add(p1)
                    
                    # Otros participantes (asumimos que son usernames)
                    for participant_name in participants_data:
                        other_user = db.query(User).filter(
                            User.username == participant_name
                        ).first()
                        
                        if other_user:
                            p = ExpenseParticipant(
                                expense_id=expense.id,
                                user_id=other_user.id,
                                amount=participant_amount
                            )
                            db.add(p)
                
                imported += 1
                
            except Exception as e:
                print_warning(f"Fila {idx}: Error - {e}")
                errors += 1
        
        db.commit()
        
        print_success(f"\n✅ Importación completada")
        print(f"  - Gastos importados: {imported}")
        print(f"  - Errores: {errors}")
        
    except Exception as e:
        print_error(f"Error durante importación: {e}")
        db.rollback()
    finally:
        db.close()

def create_template_excel(filepath: str = "gastos_template.xlsx"):
    """Crear plantilla Excel para llenar"""
    
    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Gastos"
        
        # Headers
        headers = [
            "Fecha (DD/MM/YYYY)",
            "Concepto",
            "Categoría",
            "Monto",
            "Tarjeta",
            "Participante 1",
            "Participante 2",
            "Participante 3"
        ]
        
        sheet.append(headers)
        
        # Ejemplo
        sheet.append([
            "15/01/2024",
            "Cena en restaurante",
            "Comida",
            "300.00",
            "Visa Personal",
            "juan",
            "maria",
            ""
        ])
        
        sheet.append([
            "16/01/2024",
            "Uber",
            "Transporte",
            "45.50",
            "Visa Personal",
            "",
            "",
            ""
        ])
        
        # Ajustar ancho de columnas
        for col in sheet.columns:
            max_length = 0
            for cell in col:
                max_length = max(max_length, len(str(cell.value)))
            sheet.column_dimensions[col[0].column_letter].width = max_length + 2
        
        workbook.save(filepath)
        print_success(f"Plantilla creada: {filepath}")
        print("Instrucciones:")
        print("1. Abrir la plantilla en Excel")
        print("2. Llenar tus gastos (mínimo: Fecha, Concepto, Categoría, Monto, Tarjeta)")
        print("3. Los participantes deben ser usernames de la BD")
        print("4. Ejecutar: python import_excel.py gastos_template.xlsx")
        
    except Exception as e:
        print_error(f"No se pudo crear plantilla: {e}")

def main():
    print("\n" + "="*60)
    print("IMPORTADOR DE GASTOS DESDE EXCEL")
    print("="*60)
    
    if len(sys.argv) < 2:
        # Crear template
        print("\nUso: python import_excel.py <archivo.xlsx>")
        print("\n¿Crear plantilla Excel? (s/n)")
        if input().lower() == 's':
            create_template_excel()
        return
    
    filepath = sys.argv[1]
    import_from_excel(filepath)

if __name__ == "__main__":
    main()
